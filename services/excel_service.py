#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Servizio per la gestione dei file Excel.
"""

import logging
import re
import os
from typing import Dict, Any, List, Tuple, Optional

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
    from openpyxl.utils import get_column_letter
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

from config import get_config
from models.workout import Workout, WorkoutStep, Target


class ExcelService:
    """Servizio per la gestione dei file Excel."""
    
    @staticmethod
    def check_pandas():
        """
        Verifica che pandas sia disponibile.
        
        Raises:
            ImportError: Se pandas non è disponibile
        """
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas e openpyxl sono richiesti per gestire i file Excel")
    

    @staticmethod
    def import_workouts(file_path: str) -> List[Tuple[str, Workout]]:
        """
        Importa allenamenti da un file Excel con multipli fogli.
        
        Args:
            file_path: Percorso del file Excel
            
        Returns:
            Lista di tuple (nome, allenamento)
            
        Raises:
            ImportError: Se pandas non è disponibile
            ValueError: Se il file non può essere letto
        """
        ExcelService.check_pandas()
        
        try:
            # Leggi il file Excel
            xls = pd.ExcelFile(file_path)
            
            # Ottieni la configurazione corrente
            config = get_config()
            
            # Processa i fogli di configurazione se presenti
            if 'Config' in xls.sheet_names:
                config_df = pd.read_excel(file_path, sheet_name='Config')
                
                logging.info(f"Foglio 'Config' trovato. Colonne: {config_df.columns.tolist()}")
                
                # Verifica che le colonne necessarie siano presenti
                if 'Parameter' in config_df.columns and 'Value' in config_df.columns:
                    for _, row in config_df.iterrows():
                        param = row.get('Parameter')
                        value = row.get('Value')
                        
                        if pd.notna(param) and pd.notna(value):
                            str_value = str(value).strip()
                            
                            # Imposta i parametri nella configurazione
                            if param == 'athlete_name':
                                config.set('athlete_name', str_value)
                            elif param == 'name_prefix':
                                config.set('planning.name_prefix', str_value)
                            elif param == 'race_day':
                                config.set('planning.race_day', str_value)
                            elif param == 'preferred_days':
                                # Converti la stringa in lista di interi
                                try:
                                    # Gestisci vari formati
                                    if '[' in str_value:
                                        # Formato: [1, 3, 5] o [1,3,5]
                                        days_str = str_value.strip('[]')
                                        days = [int(d.strip()) for d in days_str.split(',') if d.strip()]
                                    else:
                                        # Formato: 1,3,5 o 1 3 5
                                        days = [int(d.strip()) for d in re.split('[,\\s]+', str_value) if d.strip()]
                                    config.set('planning.preferred_days', days)
                                except Exception as e:
                                    logging.warning(f"Impossibile parsare preferred_days '{str_value}': {e}")
                            elif param.startswith('margins.'):
                                # Gestione margini
                                margin_type = param.split('.')[1]
                                if margin_type == 'faster':
                                    config.set('sports.running.margins.faster', value)
                                    config.set('sports.swimming.margins.faster', value)
                                elif margin_type == 'slower':
                                    config.set('sports.running.margins.slower', value)
                                    config.set('sports.swimming.margins.slower', value)
                                elif margin_type == 'power_up':
                                    config.set('sports.cycling.margins.power_up', value)
                                elif margin_type == 'power_down':
                                    config.set('sports.cycling.margins.power_down', value)
                                elif margin_type == 'hr_up':
                                    config.set('hr_margins.hr_up', value)
                                elif margin_type == 'hr_down':
                                    config.set('hr_margins.hr_down', value)
                                    
                    logging.info("Configurazione importata con successo")
            
            # Carica i valori di passo, nuoto e potenza se presenti
            if 'Paces' in xls.sheet_names:
                paces_df = pd.read_excel(file_path, sheet_name='Paces')
                
                logging.info(f"Foglio 'Paces' trovato. Colonne: {paces_df.columns.tolist()}")
                
                # Verifica l'esistenza delle sezioni
                sections = paces_df[paces_df['Value'].isna() & ~paces_df['Name'].isna()]['Name'].tolist()
                logging.info(f"Sezioni trovate: {sections}")
                
                run_section = "RITMI PER LA CORSA"
                cycle_section = "POTENZA PER IL CICLISMO"
                swim_section = "PASSI VASCA PER IL NUOTO"
                
                current_section = None
                
                # Dizionari per raccogliere tutti i valori di ogni sezione
                running_paces = {}
                cycling_powers = {}
                swimming_paces = {}
                
                # Processa riga per riga
                for idx, row in paces_df.iterrows():
                    name = row.get('Name')
                    value = row.get('Value')
                    
                    # Stampa ogni riga per il debug
                    logging.debug(f"Riga {idx}: name='{name}', value='{value}'")
                    
                    # Cambiamento di sezione
                    if pd.isna(value) and pd.notna(name):
                        for section_name, section_type in [(run_section, "running"), 
                                                          (cycle_section, "cycling"), 
                                                          (swim_section, "swimming")]:
                            if section_name in str(name) or str(name) == section_name:
                                current_section = section_type
                                logging.info(f"Cambiato a sezione: {current_section}")
                                break
                    
                    # Processa i valori
                    elif pd.notna(name) and pd.notna(value) and current_section:
                        name_str = str(name).strip()
                        value_str = str(value).strip()
                        
                        logging.info(f"Processando {current_section}: {name_str} = {value_str}")
                        
                        # Aggiunge al dizionario appropriato
                        if current_section == "running":
                            running_paces[name_str] = value_str
                        elif current_section == "cycling":
                            cycling_powers[name_str] = value_str
                        elif current_section == "swimming":
                            swimming_paces[name_str] = value_str
                
                # Sostituisci completamente le sezioni con i nuovi valori
                if running_paces:
                    config.replace_section('sports.running.paces', running_paces)
                    logging.info(f"Sostituita sezione running paces con {len(running_paces)} valori")
                    
                if cycling_powers:
                    config.replace_section('sports.cycling.power_values', cycling_powers)
                    logging.info(f"Sostituita sezione cycling power con {len(cycling_powers)} valori")
                    
                if swimming_paces:
                    config.replace_section('sports.swimming.paces', swimming_paces)
                    logging.info(f"Sostituita sezione swimming paces con {len(swimming_paces)} valori")
                
                logging.info("Valori di pace/potenza importati con successo")
            
            # Carica i valori di frequenza cardiaca se presenti
            if 'HeartRates' in xls.sheet_names:
                hr_df = pd.read_excel(file_path, sheet_name='HeartRates')
                
                logging.info(f"Foglio 'HeartRates' trovato. Colonne: {hr_df.columns.tolist()}")
                
                # Dizionario per raccogliere tutti i valori HR
                heart_rates = {}
                
                # Verifica che le colonne necessarie siano presenti
                if 'Name' in hr_df.columns and 'Value' in hr_df.columns:
                    for _, row in hr_df.iterrows():
                        name = row.get('Name')
                        value = row.get('Value')
                        
                        if pd.notna(name) and pd.notna(value):
                            str_value = str(value)
                            heart_rates[name] = str_value
                            logging.info(f"Raccolto heart rate: {name} = {str_value}")
                
                # Usa replace_section per sostituire completamente la sezione
                if heart_rates:
                    config.replace_section('heart_rates', heart_rates)
                    logging.info(f"Sostituita sezione heart rates con {len(heart_rates)} valori")
                
                logging.info("Valori di frequenza cardiaca importati con successo")
            
            # Salva le modifiche alla configurazione
            config.save()
            
            # Lista degli allenamenti importati
            imported_workouts = []
            
            # Importa gli allenamenti se presenti
            if 'Workouts' in xls.sheet_names:
                workouts_df = pd.read_excel(file_path, sheet_name='Workouts')
                
                logging.info(f"Foglio 'Workouts' trovato. Righe: {len(workouts_df)}")
                
                # Processa riga per riga
                for row_idx, row in workouts_df.iterrows():
                    try:
                        # Estrai informazioni di base
                        week = row.get('Week')
                        session = row.get('Session', '')
                        sport_type = str(row.get('Sport', 'running')).lower()
                        description = row.get('Description', '')
                        steps_text = row.get('Steps', '')
                        
                        # Data nel formato DD/MM/YYYY
                        date_text = row.get('Date', '')
                        workout_date = None
                        
                        if pd.notna(date_text) and date_text:
                            try:
                                # Se è già una stringa, usala direttamente
                                if isinstance(date_text, str):
                                    date_parts = date_text.split('/')
                                    if len(date_parts) == 3:
                                        day, month, year = date_parts
                                        workout_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                                # Se è un datetime di pandas
                                else:
                                    workout_date = date_text.strftime('%Y-%m-%d')
                            except Exception as e:
                                logging.warning(f"Impossibile parsare la data '{date_text}': {e}")
                        
                        # Nome dell'allenamento - GESTIONE CORRETTA DEI TIPI
                        if pd.notna(session) and session != '':
                            # Converti session in stringa per sicurezza
                            session_str = str(session)
                            
                            # Se session contiene già il formato completo (es. "W9D5"), usalo
                            if session_str.startswith('W') and 'D' in session_str:
                                workout_name = f"{session_str} - {description}"
                            # Altrimenti, se abbiamo week, ricostruisci il nome
                            elif pd.notna(week):
                                try:
                                    # Gestisci week che può essere numero o stringa
                                    if isinstance(week, str) and week.startswith('W'):
                                        week_num = int(week[1:])
                                    else:
                                        week_num = int(week)
                                    
                                    # Gestisci session che può essere numero o stringa
                                    if isinstance(session, (int, float)):
                                        # Session è già un numero
                                        session_num = int(session)
                                        workout_name = f"W{week_num}D{session_num} - {description}"
                                    elif session_str.isdigit():
                                        # Session è una stringa numerica
                                        session_num = int(session_str)
                                        workout_name = f"W{week_num}D{session_num} - {description}"
                                    elif session_str.startswith('D'):
                                        # Session è nel formato "D5"
                                        workout_name = f"W{week_num}{session_str} - {description}"
                                    else:
                                        # Altri formati
                                        workout_name = f"{session_str} - {description}"
                                except Exception as e:
                                    logging.warning(f"Errore nel parsing week/session: {e}")
                                    workout_name = f"{session_str} - {description}"
                            else:
                                workout_name = f"{session_str} - {description}"
                        else:
                            workout_name = description
                        
                        # Log per debug
                        logging.debug(f"Week: {week} (type: {type(week)}), Session: {session} (type: {type(session)})")
                        logging.debug(f"Nome allenamento costruito: {workout_name}")

                        
                        # Creare l'allenamento
                        workout = Workout(sport_type, workout_name, description)
                        
                        # Se c'è una data, aggiungi uno step speciale con la data
                        if workout_date:
                            date_step = WorkoutStep(0, "warmup")
                            date_step.date = workout_date
                            workout.add_step(date_step)
                        
                        # Verifica che gli steps siano validi
                        if not pd.isna(steps_text) and steps_text:
                            # Parsa gli step
                            logging.info(f"Parsing steps per '{workout_name}'")
                            steps_lines = str(steps_text).strip().split('\n')
                            
                            # Tieni traccia dello stato corrente del parsing
                            current_repeat = None
                            indent_level = 0
                            
                            for line_num, line in enumerate(steps_lines, 1):
                                # Salta linee vuote
                                if not line.strip():
                                    continue
                                
                                logging.debug(f"Processing line {line_num}: {line}")
                                
                                # Calcola il livello di indentazione
                                line_indent = len(line) - len(line.lstrip())
                                
                                # Normalizza la linea
                                line = line.strip()
                                

                                # Controlla se è una ripetizione
                                if line.startswith('repeat'):
                                    try:
                                        # Rimuovi i due punti alla fine se presenti
                                        line_clean = line.replace(':', '')
                                        
                                        # Estrai il numero di ripetizioni
                                        iterations = 1  # default
                                        # Usa una regex per estrarre il numero
                                        import re
                                        numbers = re.findall(r'\d+', line_clean)
                                        if numbers:
                                            iterations = int(numbers[0])
                                        else:
                                            logging.warning(f"Nessun numero trovato in '{line}', usando 1 ripetizione")
                                        
                                        logging.debug(f"Creato repeat step con {iterations} ripetizioni")
                                        current_repeat = WorkoutStep(
                                            order=0,
                                            step_type='repeat',
                                            end_condition='iterations',
                                            end_condition_value=iterations
                                        )
                                        
                                        # Aggiungi all'allenamento
                                        workout.add_step(current_repeat)
                                        indent_level = line_indent
                                        
                                    except Exception as e:
                                        logging.error(f"Errore nel parsing del repeat '{line}': {str(e)}")
                                        # NON c'è continue qui, continua con il resto del codice
                                
                                # Controlla se è la fine di una ripetizione
                                elif current_repeat and line_indent <= indent_level:
                                    # Se siamo a un livello di indentazione uguale o inferiore a quello del repeat,
                                    # significa che siamo usciti dal blocco repeat
                                    logging.debug(f"Fine ripetizione (indent {line_indent} <= {indent_level})")
                                    
                                    # Salviamo l'indentazione corrente per questo step non in repeat
                                    current_line_indent = line_indent
                                    
                                    # Chiudiamo il repeat
                                    current_repeat = None
                                    indent_level = 0
                                    
                                    # NON processare questo step qui, lascia che venga processato nel blocco else sotto
                                
                                # Se non è un repeat e non è la fine di un repeat, è uno step normale
                                if not line.startswith('repeat'):
                                    try:
                                        # Determina il tipo di step
                                        if ':' not in line:
                                            logging.warning(f"Linea '{line}' non contiene ':' - saltata")
                                            continue
                                        
                                        step_type, step_data = line.split(':', 1)
                                        step_type = step_type.strip()
                                        step_data = step_data.strip()
                                        
                                        logging.debug(f"Tipo di step: {step_type}, Dati: {step_data}")
                                        
                                        # Parsa i dati dello step
                                        end_condition = "lap.button"
                                        end_condition_value = None
                                        description = ""
                                        target = None
                                        
                                        # Estrai la descrizione se presente
                                        if '--' in step_data:
                                            step_data, description = step_data.split('--', 1)
                                            step_data = step_data.strip()
                                            description = description.strip()
                                        
                                        # Estrai il target se presente
                                        if '@' in step_data:
                                            step_data, target_data = step_data.split('@', 1)
                                            step_data = step_data.strip()
                                            target_data = target_data.strip()
                                            
                                            # Determina il tipo di target
                                            if target_data.startswith('Z') and '_HR' in target_data:
                                                # Zona di frequenza cardiaca
                                                target_type = "heart.rate.zone"
                                                
                                                # Ottieni i valori HR dalla configurazione
                                                heart_rates = config.get('heart_rates', {})
                                                max_hr = int(heart_rates.get('max_hr', 180))
                                                
                                                # Cerca la zona corrispondente
                                                if target_data in heart_rates:
                                                    hr_range = heart_rates[target_data]
                                                    
                                                    if '-' in hr_range and 'max_hr' in hr_range:
                                                        # Formato: 62-76% max_hr
                                                        parts = hr_range.split('-')
                                                        min_percent = float(parts[0])
                                                        max_percent = float(parts[1].split('%')[0])
                                                        
                                                        # Converti in valori assoluti
                                                        target_from = int(min_percent * max_hr / 100)
                                                        target_to = int(max_percent * max_hr / 100)
                                                        target = Target(target_type, target_to, target_from)
                                                        target.target_zone_name = target_data
                                                    else:
                                                        # Default se formato non riconosciuto
                                                        target = Target(target_type, 140, 120)
                                                        target.target_zone_name = target_data
                                                else:
                                                    # Default se la zona non è trovata
                                                    target = Target(target_type, 140, 120)
                                            
                                            elif target_data.startswith('Z') or target_data in ['recovery', 'threshold', 'marathon', 'race_pace']:
                                                # Zona di passo
                                                target_type = "pace.zone"
                                                
                                                # Ottieni i valori di passo dalla configurazione
                                                paces = config.get('sports.running.paces', {})
                                                
                                                # Cerca la zona corrispondente
                                                if target_data in paces:
                                                    pace_range = paces[target_data]
                                                    
                                                    if '-' in pace_range:
                                                        # Formato: min:sec-min:sec
                                                        min_pace, max_pace = pace_range.split('-')
                                                        
                                                        # Converti da min:sec a secondi
                                                        def parse_pace(pace_str):
                                                            parts = pace_str.strip().split(':')
                                                            return int(parts[0]) * 60 + int(parts[1])
                                                        
                                                        try:
                                                            min_pace_secs = parse_pace(min_pace)
                                                            max_pace_secs = parse_pace(max_pace)
                                                            
                                                            # Converti da secondi a m/s (inverti min e max)
                                                            target_from = 1000 / max_pace_secs  # Passo più veloce
                                                            target_to = 1000 / min_pace_secs    # Passo più lento
                                                            target = Target(target_type, target_to, target_from)
                                                            target.target_zone_name = target_data
                                                        except (ValueError, IndexError):
                                                            # Default se non riesce a interpretare
                                                            target = Target(target_type, 2.5, 3.0)
                                                            target.target_zone_name = target_data
                                                    else:
                                                        # Valore singolo
                                                        try:
                                                            pace_parts = pace_range.strip().split(':')
                                                            pace_secs = int(pace_parts[0]) * 60 + int(pace_parts[1])
                                                            # Per un valore singolo, usa lo stesso valore per from e to
                                                            pace_ms = 1000 / pace_secs
                                                            target = Target(target_type, pace_ms, pace_ms)
                                                            target.target_zone_name = target_data
                                                        except (ValueError, IndexError):
                                                            # Default se non riesce a interpretare
                                                            target = Target(target_type, 3.0, 3.0)
                                                            target.target_zone_name = target_data
                                                else:
                                                    # Default se la zona non è trovata
                                                    target = Target(target_type, 2.5, 3.0)
                                            
                                            elif ':' in target_data:
                                                # Passo specifico (es. "6:00" o "5:00-5:30")
                                                target_type = "pace.zone"
                                                
                                                if '-' in target_data:
                                                    # Formato: min:sec-min:sec
                                                    min_pace, max_pace = target_data.split('-')
                                                    
                                                    # Converti da min:sec a secondi
                                                    def parse_pace(pace_str):
                                                        parts = pace_str.strip().split(':')
                                                        return int(parts[0]) * 60 + int(parts[1])
                                                    
                                                    try:
                                                        min_pace_secs = parse_pace(min_pace)
                                                        max_pace_secs = parse_pace(max_pace)
                                                        
                                                        # Converti da secondi a m/s (inverti min e max)
                                                        target_from = 1000 / max_pace_secs  # Passo più veloce
                                                        target_to = 1000 / min_pace_secs    # Passo più lento
                                                        target = Target(target_type, target_to, target_from)
                                                    except (ValueError, IndexError):
                                                        # Default se non riesce a interpretare
                                                        target = Target(target_type, 2.5, 3.0)
                                                else:
                                                    # Formato: min:sec (passo singolo)
                                                    try:
                                                        parts = target_data.strip().split(':')
                                                        pace_secs = int(parts[0]) * 60 + int(parts[1])
                                                        
                                                        # Per un passo singolo, usa lo stesso valore per from e to
                                                        pace_ms = 1000 / pace_secs
                                                        target = Target(target_type, pace_ms, pace_ms)
                                                    except (ValueError, IndexError):
                                                        # Default se non riesce a interpretare
                                                        target = Target(target_type, 3.0, 3.0)
                                            
                                            else:
                                                # Altri tipi di target...
                                                target = Target("no.target", None, None)
                                        
                                        # Parsa la durata/distanza
                                        if step_data == 'lap-button':
                                            end_condition = 'lap.button'
                                        elif ':' in step_data and 'min' in step_data:
                                            # Minuti con secondi: 5:30min
                                            end_condition = 'time'
                                            try:
                                                time_part = step_data.replace('min', '').strip()
                                                if ':' in time_part:
                                                    mins, secs = time_part.split(':')
                                                    end_condition_value = int(mins) * 60 + int(secs)
                                                    logging.debug(f"Tempo parsato: {mins}:{secs}min = {end_condition_value}s")
                                                else:
                                                    # Solo minuti
                                                    mins = int(time_part)
                                                    end_condition_value = mins * 60
                                                    logging.debug(f"Tempo parsato: {mins}min = {end_condition_value}s")
                                            except Exception as e:
                                                end_condition_value = 60
                                                logging.warning(f"Errore nel parsing del tempo '{step_data}': {e}")
                                        elif 'min' in step_data:
                                            # Solo minuti: 10min
                                            end_condition = 'time'
                                            try:
                                                time_part = step_data.replace('min', '').strip()
                                                mins = int(time_part)
                                                end_condition_value = mins * 60
                                                logging.debug(f"Tempo parsato: {mins}min = {end_condition_value}s")
                                            except Exception as e:
                                                end_condition_value = 60
                                                logging.warning(f"Errore nel parsing del tempo '{time_part}': {e}")
                                        elif step_data.strip().endswith('m') and not step_data.strip().endswith('km'):
                                            # Metri
                                            end_condition = 'distance'
                                            try:
                                                # Rimuovi 'm' alla fine
                                                distance_part = step_data.split(' ')[0].replace('m', '')
                                                distance = float(distance_part.strip())
                                                end_condition_value = distance
                                                logging.debug(f"Distanza parsata: {distance}m")
                                            except Exception as e:
                                                end_condition_value = 100
                                                logging.warning(f"Errore nel parsing della distanza '{step_data}': {e}")
                                        elif 'km' in step_data:
                                            end_condition = 'distance'
                                            try:
                                                # Rimuovi 'km' alla fine
                                                distance_part = step_data.split(' ')[0].replace('km', '')
                                                distance = float(distance_part.strip())
                                                end_condition_value = distance * 1000  # Converti in metri
                                                logging.debug(f"Distanza parsata: {distance}km = {end_condition_value}m")
                                            except Exception as e:
                                                end_condition_value = 1000
                                                logging.warning(f"Errore nel parsing della distanza '{step_data}': {e}")
                                        
                                        # Crea lo step
                                        step = WorkoutStep(
                                            order=0,
                                            step_type=step_type,
                                            description=description,
                                            end_condition=end_condition,
                                            end_condition_value=end_condition_value,
                                            target=target
                                        )
                                        
                                        logging.debug(f"Creato step: {step}")
                                        
                                        # Aggiungi lo step al gruppo corrente o all'allenamento
                                        if current_repeat and line_indent > indent_level:
                                            current_repeat.add_step(step)
                                            logging.debug(f"Step aggiunto al repeat")
                                        else:
                                            workout.add_step(step)
                                            logging.debug(f"Step aggiunto all'allenamento")
                                        
                                    except Exception as e:
                                        logging.error(f"Errore nel parsing dello step '{line}': {str(e)}")
                                        continue

                        
                        # Aggiungi l'allenamento alla lista degli importati
                        imported_workouts.append((workout_name, workout))
                        logging.info(f"Allenamento '{workout_name}' importato con successo, con {len(workout.workout_steps)} step")
                        
                    except Exception as e:
                        logging.error(f"Errore nell'importazione dell'allenamento alla riga {row_idx}: {str(e)}")
                        raise ValueError(f"Errore nell'importazione dell'allenamento alla riga {row_idx}: {str(e)}")
            
            # Log del risultato
            logging.info(f"Importati {len(imported_workouts)} allenamenti")
            return imported_workouts
            
        except Exception as e:
            logging.error(f"Errore nell'importazione degli allenamenti da Excel: {str(e)}")
            raise


    @staticmethod
    def export_workouts(workouts: List[Tuple[str, Workout]], file_path: str, custom_config: Optional[Dict[str, Any]] = None) -> None:
        """
        Esporta allenamenti in un file Excel con multipli fogli.
        
        Args:
            workouts: Lista di tuple (nome, allenamento)
            file_path: Percorso del file
            custom_config: Configurazione personalizzata (opzionale)
            
        Raises:
            ImportError: Se pandas non è disponibile
            IOError: Se il file non può essere scritto
        """
        # Verifica che pandas sia disponibile
        ExcelService.check_pandas()
        
        try:
            # Ottieni la configurazione
            config = get_config()
            
            # Crea i DataFrame per ogni foglio
            config_df = pd.DataFrame(columns=['Parametro', 'Valore', 'Descrizione'])
            paces_df = pd.DataFrame(columns=['Name', 'Value', 'Note'])
            heart_rates_df = pd.DataFrame(columns=['Name', 'Value', 'Description'])
            workouts_df = pd.DataFrame(columns=['Week', 'Session', 'Date', 'Sport', 'Description', 'Steps'])
            examples_df = pd.DataFrame(columns=['Type', 'Example', 'Description'])
                
            # Popola il DataFrame della configurazione
            # Ottieni la data della gara nel formato YYYY-MM-DD
            race_day = config.get('planning.race_day', '')
            # Converti nel formato GG/MM/YYYY se necessario
            if race_day and '-' in race_day:
                try:
                    year, month, day = race_day.split('-')
                    race_day_display = f"{day}/{month}/{year}"
                except:
                    race_day_display = race_day
            else:
                race_day_display = race_day
            
            config_rows = [
                {'Parametro': 'athlete_name', 'Valore': config.get('athlete_name', ''), 
                 'Descrizione': 'Nome dell\'atleta'},
                {'Parametro': 'name_prefix', 'Valore': config.get('planning.name_prefix', ''), 
                 'Descrizione': 'Prefisso per i nomi degli allenamenti'},
                {'Parametro': 'race_day', 'Valore': race_day_display, 
                 'Descrizione': 'Data della gara (GG/MM/AAAA)'},
                {'Parametro': 'preferred_days', 'Valore': str(config.get('planning.preferred_days', [1, 3, 5])), 
                 'Descrizione': 'Giorni preferiti per gli allenamenti [0=Lunedì, 6=Domenica]'},
                {'Parametro': 'margin.faster', 'Valore': config.get('sports.running.margins.faster', '0:05'), 
                 'Descrizione': 'Margine più veloce per corsa/nuoto'},
                {'Parametro': 'margin.slower', 'Valore': config.get('sports.running.margins.slower', '0:05'), 
                 'Descrizione': 'Margine più lento per corsa/nuoto'},
                {'Parametro': 'margin.power_up', 'Valore': config.get('sports.cycling.margins.power_up', 10), 
                 'Descrizione': 'Margine potenza superiore per ciclismo'},
                {'Parametro': 'margin.power_down', 'Valore': config.get('sports.cycling.margins.power_down', 10), 
                 'Descrizione': 'Margine potenza inferiore per ciclismo'},
                {'Parametro': 'margin.hr_up', 'Valore': config.get('hr_margins.hr_up', 5), 
                 'Descrizione': 'Margine superiore frequenza cardiaca'},
                {'Parametro': 'margin.hr_down', 'Valore': config.get('hr_margins.hr_down', 5), 
                 'Descrizione': 'Margine inferiore frequenza cardiaca'},
            ]
            config_df = pd.DataFrame(config_rows)
            
            # Popola il DataFrame dei paces
            paces_rows = []
            
            # Aggiungi titolo per la sezione dei ritmi di corsa
            paces_rows.append({'Name': 'RITMI PER LA CORSA (min/km)', 'Value': None, 'Note': None})
            
            # Aggiungi i ritmi per la corsa
            running_paces = config.get('sports.running.paces', {})
            for name, value in running_paces.items():
                description = ""
                if name == 'Z1':
                    description = "Ritmo facile (zona 1, recupero)"
                elif name == 'Z2':
                    description = "Ritmo aerobico (zona 2, endurance)"
                elif name == 'Z3':
                    description = "Ritmo medio (zona 3, soglia aerobica)"
                elif name == 'Z4':
                    description = "Ritmo soglia (zona 4, soglia anaerobica)"
                elif name == 'Z5':
                    description = "Ritmo VO2max (zona 5, anaerobico)"
                elif name == 'recovery':
                    description = "Ritmo recupero (molto lento)"
                elif name == 'threshold':
                    description = "Ritmo soglia personalizzato"
                elif name == 'marathon':
                    description = "Ritmo maratona personalizzato"
                elif name == 'race_pace':
                    description = "Ritmo gara personalizzato"
                
                paces_rows.append({'Name': name, 'Value': value, 'Note': description})
            
            # Aggiungi una riga vuota per separazione
            paces_rows.append({'Name': None, 'Value': None, 'Note': None})
            
            # Aggiungi titolo per la sezione potenza ciclismo
            paces_rows.append({'Name': 'POTENZA PER IL CICLISMO (Watt)', 'Value': None, 'Note': None})
            
            # Aggiungi i valori di potenza per il ciclismo
            power_values = config.get('sports.cycling.power_values', {})
            for name, value in power_values.items():
                description = ""
                if name == 'ftp':
                    description = "Functional Threshold Power (W)"
                elif name == 'Z1':
                    description = "Recupero attivo (55-70% FTP)"
                elif name == 'Z2':
                    description = "Endurance (70-85% FTP)"
                elif name == 'Z3':
                    description = "Tempo/Soglia (85-100% FTP)"
                elif name == 'Z4':
                    description = "VO2max (100-120% FTP)"
                elif name == 'Z5':
                    description = "Capacità anaerobica (120-150% FTP)"
                elif name == 'Z6':
                    description = "Potenza neuromuscolare (>150% FTP)"
                elif name == 'recovery':
                    description = "Recupero (<55% FTP)"
                elif name == 'threshold':
                    description = "Soglia (94-106% FTP)"
                elif name == 'sweet_spot':
                    description = "Sweet Spot (88-94% FTP)"
                
                paces_rows.append({'Name': name, 'Value': value, 'Note': description})
            
            # Aggiungi una riga vuota per separazione
            paces_rows.append({'Name': None, 'Value': None, 'Note': None})
            
            # Aggiungi titolo per la sezione passi vasca
            paces_rows.append({'Name': 'PASSI VASCA PER IL NUOTO (min/100m)', 'Value': None, 'Note': None})
            
            # Aggiungi i passi vasca per il nuoto
            swim_paces = config.get('sports.swimming.paces', {})
            for name, value in swim_paces.items():
                description = ""
                if name == 'Z1':
                    description = "Ritmo facile (zona 1)"
                elif name == 'Z2':
                    description = "Ritmo aerobico (zona 2)"
                elif name == 'Z3':
                    description = "Ritmo medio (zona 3)"
                elif name == 'Z4':
                    description = "Ritmo soglia (zona 4)"
                elif name == 'Z5':
                    description = "Ritmo VO2max (zona 5)"
                elif name == 'recovery':
                    description = "Ritmo recupero"
                elif name == 'threshold':
                    description = "Ritmo soglia personalizzato"
                elif name == 'sprint':
                    description = "Ritmo sprint"
                
                paces_rows.append({'Name': name, 'Value': value, 'Note': description})
            
            # Crea il DataFrame dai dati raccolti
            paces_df = pd.DataFrame(paces_rows)
            
            # Popola il DataFrame delle frequenze cardiache
            hr_rows = []
            heart_rates = config.get('heart_rates', {})
            for name, value in heart_rates.items():
                description = ""
                if name == 'max_hr':
                    description = "Frequenza cardiaca massima"
                elif name == 'rest_hr':
                    description = "Frequenza cardiaca a riposo"
                elif name == 'Z1_HR':
                    description = "Zona 1 (recupero attivo)"
                elif name == 'Z2_HR':
                    description = "Zona 2 (fondo facile)"
                elif name == 'Z3_HR':
                    description = "Zona 3 (fondo medio)"
                elif name == 'Z4_HR':
                    description = "Zona 4 (soglia anaerobica)"
                elif name == 'Z5_HR':
                    description = "Zona 5 (massimale)"
                
                hr_rows.append({'Name': name, 'Value': value, 'Description': description})
            
            # Crea il DataFrame dai dati raccolti
            heart_rates_df = pd.DataFrame(hr_rows)
            
            # Popola il DataFrame degli esempi
            examples_rows = [
                {'Type': 'CONFIGURAZIONE', 'Example': '', 'Description': 'Esempi di configurazione'},
                {'Type': 'Preferred Days', 'Example': '[1, 3, 5]', 'Description': 'Giorni preferiti: Lunedi=0, Martedi=1, ...'},
                {'Type': 'Race Day', 'Example': '2025-06-15', 'Description': 'Data della gara in formato YYYY-MM-DD'},
                {'Type': '', 'Example': '', 'Description': ''},
                {'Type': 'STEP ALLENAMENTO', 'Example': '', 'Description': 'Esempi di step degli allenamenti'},
                {'Type': 'Riscaldamento', 'Example': 'warmup: 10min @ Z2', 'Description': 'Riscaldamento di 10 minuti in Zona 2'},
                {'Type': 'Intervallo', 'Example': 'interval: 400m @ Z4', 'Description': 'Intervallo di 400m in Zona 4'},
                {'Type': 'Recupero', 'Example': 'recovery: 2min @ Z1', 'Description': 'Recupero di 2 minuti in Zona 1'},
                {'Type': 'Defaticamento', 'Example': 'cooldown: 5min @ Z1_HR', 'Description': 'Defaticamento di 5 min in Zona HR 1'},
                {'Type': 'Repeat', 'Example': 'repeat 4:', 'Description': 'Ripeti 4 volte gli step indentati'},
                {'Type': 'Descrizione', 'Example': 'interval: 1km @ Z3 -- Run Strong', 'Description': 'Aggiungere descrizione dopo --'},
                {'Type': '', 'Example': '', 'Description': ''},
                {'Type': 'TARGET', 'Example': '', 'Description': 'Esempi di target'},
                {'Type': 'Pace Zone', 'Example': '@ Z2', 'Description': 'Target pace in Zona 2'},
                {'Type': 'Heart Rate', 'Example': '@ Z2_HR', 'Description': 'Target frequenza cardiaca in Zona 2'},
                {'Type': 'Power', 'Example': '@ threshold', 'Description': 'Target potenza a soglia'},
                {'Type': 'Specifico HR', 'Example': '@ 140-150 bpm', 'Description': 'Target frequenza specifica'},
                {'Type': 'Specifico Pace', 'Example': '@ 5:00-5:30', 'Description': 'Target pace specifico (min:sec per km)'},
                {'Type': 'Specifico Power', 'Example': '@ 250W', 'Description': 'Target potenza specifica (Watt)'},
                {'Type': '', 'Example': '', 'Description': ''},
                {'Type': 'DURATA', 'Example': '', 'Description': 'Esempi di durata'},
                {'Type': 'Tempo', 'Example': '10min', 'Description': '10 minuti'},
                {'Type': 'Tempo con sec', 'Example': '5:30min', 'Description': '5 minuti e 30 secondi'},
                {'Type': 'Distanza', 'Example': '400m', 'Description': '400 metri'},
                {'Type': 'Distanza km', 'Example': '5km', 'Description': '5 chilometri'},
                {'Type': 'Lap Button', 'Example': 'lap-button', 'Description': 'Terminato manualmente'},
            ]
            examples_df = pd.DataFrame(examples_rows)
            
            # Popola il DataFrame degli allenamenti
            workout_rows = []
            
            for name, workout in workouts:
                # Estrai informazioni dal nome (se disponibile)
                week = None
                session = None
                description = workout.workout_name
                
                if workout.workout_name.startswith('W') and 'D' in workout.workout_name:
                    try:
                        parts = workout.workout_name.split('D')
                        week_part = parts[0].strip()
                        week = int(week_part[1:])
                        
                        rest_parts = parts[1].split('-', 1)
                        session = int(rest_parts[0].strip())
                        
                        if len(rest_parts) > 1:
                            description = rest_parts[1].strip()
                    except:
                        pass
                        
                # Trova la data dell'allenamento (se presente)
                workout_date = None
                for step in workout.workout_steps:
                    if hasattr(step, 'date') and step.date:
                        workout_date = step.date
                        break
                
                # Formatta la data se esiste (DD/MM/YYYY)
                formatted_date = None
                if workout_date:
                    try:
                        # Parse della data e riformattazione
                        date_parts = workout_date.split('-')
                        if len(date_parts) == 3:
                            year, month, day = date_parts
                            formatted_date = f"{day}/{month}/{year}"
                        else:
                            formatted_date = workout_date  # Mantiene originale se formato errato
                    except:
                        formatted_date = workout_date  # Mantiene originale in caso di errore
                
                # Converti gli step in formato testo
                steps_text = ExcelService.format_steps_for_export(workout)
                
                # Aggiungi al DataFrame - Nota l'ordine cambiato: Session prima di Date
                workout_rows.append({
                    'Week': week,
                    'Session': session,  # Ora prima di Date
                    'Date': formatted_date,  # Ora dopo Session con data formattata
                    'Sport': workout.sport_type.capitalize(),
                    'Description': description,
                    'Steps': steps_text
                })
            
            # Crea il DataFrame dai dati raccolti
            workouts_df = pd.DataFrame(workout_rows)
            
            # Salva i DataFrame in un file Excel
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Esporta i dati in fogli Excel
                config_df.to_excel(writer, sheet_name='Config', index=False)
                paces_df.to_excel(writer, sheet_name='Paces', index=False)
                heart_rates_df.to_excel(writer, sheet_name='HeartRates', index=False)
                workouts_df.to_excel(writer, sheet_name='Workouts', index=False)
                examples_df.to_excel(writer, sheet_name='Examples', index=False)
                
                # Ottieni il workbook e i worksheets
                workbook = writer.book
                
                # Definizione degli stili
                # Stile per intestazioni
                header_style = NamedStyle(name="header_style")
                header_style.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
                header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                header_style.border = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000")
                )
                
                # Stile per celle di dati normali
                normal_style = NamedStyle(name="normal_style")
                normal_style.font = Font(name='Arial', size=11)
                normal_style.alignment = Alignment(wrap_text=True, vertical="top")
                normal_style.border = Border(
                    left=Side(border_style="thin", color="D9D9D9"),
                    right=Side(border_style="thin", color="D9D9D9"),
                    top=Side(border_style="thin", color="D9D9D9"),
                    bottom=Side(border_style="thin", color="D9D9D9")
                )
                
                # Stile per celle di dati alternate (righe dispari)
                alt_style = NamedStyle(name="alt_style")
                alt_style.font = Font(name='Arial', size=11)
                alt_style.fill = PatternFill(start_color="EBF1F9", end_color="EBF1F9", fill_type="solid")
                alt_style.alignment = Alignment(wrap_text=True, vertical="top")
                alt_style.border = Border(
                    left=Side(border_style="thin", color="D9D9D9"),
                    right=Side(border_style="thin", color="D9D9D9"),
                    top=Side(border_style="thin", color="D9D9D9"),
                    bottom=Side(border_style="thin", color="D9D9D9")
                )
                
                # Stile per titoli di sezione (usato in Paces)
                section_style = NamedStyle(name="section_style")
                section_style.font = Font(name='Arial', size=12, bold=True)
                section_style.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
                section_style.alignment = Alignment(horizontal="left", vertical="center")
                section_style.border = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000")
                )
                
                # Aggiungi gli stili al workbook
                if "header_style" not in workbook.named_styles:
                    workbook.add_named_style(header_style)
                if "normal_style" not in workbook.named_styles:
                    workbook.add_named_style(normal_style)
                if "alt_style" not in workbook.named_styles:
                    workbook.add_named_style(alt_style)
                if "section_style" not in workbook.named_styles:
                    workbook.add_named_style(section_style)
                
                # Funzione per applicare stili a un foglio
                def style_worksheet(worksheet, col_widths, freeze_panes="A2"):
                    # Imposta larghezza colonne
                    for col_letter, width in col_widths.items():
                        worksheet.column_dimensions[col_letter].width = width
                    
                    # Applica stile alle intestazioni
                    for cell in worksheet[1]:
                        cell.style = "header_style"
                        
                    # Imposta altezza intestazioni
                    worksheet.row_dimensions[1].height = 30
                    
                    # Applica stili alternati alle righe di dati
                    max_row = worksheet.max_row
                    for row in range(2, max_row + 1):
                        for cell in worksheet[row]:
                            if row % 2 == 0:  # Righe pari
                                cell.style = "normal_style"
                            else:  # Righe dispari
                                cell.style = "alt_style"
                    
                    # Freezing panes (blocca intestazioni)
                    if freeze_panes:
                        worksheet.freeze_panes = freeze_panes
                        
                    # Aggiungi filtro automatico alle intestazioni
                    worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
                    
                    # Regola l'altezza delle righe in base al contenuto
                    for row in range(2, max_row + 1):
                        # Imposta un'altezza minima ragionevole
                        row_height = 20  # Altezza minima
                        
                        # Se ci sono celle con molto testo, aumenta l'altezza
                        for cell in worksheet[row]:
                            if cell.value and isinstance(cell.value, str) and len(cell.value) > 100:
                                row_height = max(row_height, 75)  # Celle con molto testo
                            elif cell.value and isinstance(cell.value, str) and len(cell.value) > 50:
                                row_height = max(row_height, 40)  # Celle con testo medio
                        
                        worksheet.row_dimensions[row].height = row_height
                
                # Formatta il foglio Config
                config_sheet = writer.sheets['Config']
                config_col_widths = {'A': 30, 'B': 35, 'C': 45}
                style_worksheet(config_sheet, config_col_widths)
                
                # Assicura che la colonna Valore in Config sia formattata come testo
                for row in range(2, len(config_df) + 2):
                    cell = config_sheet.cell(row=row, column=2)  # Colonna B (Valore)
                    cell.number_format = '@'  # Formato testo
                    if cell.value is not None:
                        cell.value = str(cell.value)
                
                # Formatta il foglio Paces
                paces_sheet = writer.sheets['Paces']
                paces_col_widths = {'A': 30, 'B': 25, 'C': 45}
                style_worksheet(paces_sheet, paces_col_widths)
                
                # Applica stile speciale alle righe di intestazioni di sezione in Paces
                run_section = "RITMI PER LA CORSA"
                cycle_section = "POTENZA PER IL CICLISMO"
                swim_section = "PASSI VASCA PER IL NUOTO"
                
                for row in range(1, paces_sheet.max_row + 1):
                    cell_val = paces_sheet.cell(row=row, column=1).value
                    if cell_val and isinstance(cell_val, str) and (run_section in cell_val or cycle_section in cell_val or swim_section in cell_val):
                        for cell in paces_sheet[row]:
                            cell.style = "section_style"
                        paces_sheet.row_dimensions[row].height = 25
                
                # Assicura che la colonna Value in Paces sia formattata come testo
                for row in range(2, len(paces_df) + 2):
                    cell = paces_sheet.cell(row=row, column=2)  # Colonna B (Value)
                    cell.number_format = '@'  # Formato testo
                    if cell.value is not None:
                        cell.value = str(cell.value)
                
                # Formatta il foglio HeartRates
                hr_sheet = writer.sheets['HeartRates']
                hr_col_widths = {'A': 30, 'B': 25, 'C': 45}
                style_worksheet(hr_sheet, hr_col_widths)
                
                # Assicura che la colonna Value in HeartRates sia formattata come testo
                for row in range(2, len(heart_rates_df) + 2):
                    cell = hr_sheet.cell(row=row, column=2)  # Colonna B (Value)
                    cell.number_format = '@'  # Formato testo
                    if cell.value is not None:
                        cell.value = str(cell.value)
                
                # Formatta il foglio Workouts
                workouts_sheet = writer.sheets['Workouts']
                workouts_col_widths = {'A': 15, 'B': 15, 'C': 15, 'D': 20, 'E': 35, 'F': 90}
                style_worksheet(workouts_sheet, workouts_col_widths)
                
                # Imposta altezza maggiore per le righe nel foglio Workouts
                for row in range(2, len(workouts_df) + 2):
                    workouts_sheet.row_dimensions[row].height = 150
                
                # Formatta il foglio Examples
                examples_sheet = writer.sheets['Examples']
                examples_col_widths = {'A': 30, 'B': 35, 'C': 50}
                style_worksheet(examples_sheet, examples_col_widths)
                
                # Applica stile speciale alle righe di intestazioni di sezione in Examples
                for row in range(1, examples_sheet.max_row + 1):
                    cell_val = examples_sheet.cell(row=row, column=1).value
                    cell_val_col2 = examples_sheet.cell(row=row, column=2).value
                    
                    # Identifica le righe di intestazione (hanno colonna B vuota e colonna A non vuota)
                    if cell_val and isinstance(cell_val, str) and cell_val.isupper() and (cell_val_col2 is None or cell_val_col2 == ''):
                        for cell in examples_sheet[row]:
                            cell.style = "section_style"
                        examples_sheet.row_dimensions[row].height = 25
            
        except Exception as e:
            logging.error(f"Errore nell'esportazione degli allenamenti in Excel: {str(e)}")
            raise

    @staticmethod
    def format_steps_for_export(workout: Workout, indent_level: int = 0) -> str:
        """
        Formatta gli step di un allenamento per l'esportazione in Excel.
        
        Args:
            workout: Allenamento da formattare
            indent_level: Livello di indentazione corrente
            
        Returns:
            Testo formattato degli step
        """
        steps_text = []
        
        # Aggiungi gli step (salta gli step con la data)
        steps_to_add = [s for s in workout.workout_steps if not (hasattr(s, 'date') and s.date)]
        
        for step in steps_to_add:
            # Indentazione
            indent = "    " * indent_level  # Uso 4 spazi per l'indentazione per maggiore chiarezza
            
            # Caso speciale: repeat
            if step.step_type == 'repeat':
                steps_text.append(f"{indent}repeat {step.end_condition_value}:")
                
                # Aggiungi gli step figli con indentazione aumentata
                for child_step in step.workout_steps:
                    child_text = ExcelService.format_step_for_export(child_step, indent_level + 1)
                    steps_text.append(child_text)
            else:
                # Step normale
                step_text = ExcelService.format_step_for_export(step, indent_level)
                steps_text.append(step_text)
        
        return "\n".join(steps_text)


    @staticmethod
    def format_step_for_export(step: WorkoutStep, indent_level: int = 0) -> str:
        """
        Formatta un singolo step per l'esportazione in Excel.
        
        Args:
            step: Step da formattare
            indent_level: Livello di indentazione
            
        Returns:
            Testo formattato dello step
        """
        # Indentazione
        indent = "    " * indent_level  # Uso 4 spazi per l'indentazione per maggiore chiarezza
        
        # Nome del tipo di step
        step_type = step.step_type
        
        # Valore della durata
        duration = ""
        
        if step.end_condition == 'lap.button':
            duration = "lap-button"
        elif step.end_condition == 'time':
            # Formatta il tempo
            if isinstance(step.end_condition_value, (int, float)):
                seconds = int(step.end_condition_value)
                minutes = seconds // 60
                seconds = seconds % 60
                
                if seconds > 0:
                    duration = f"{minutes}:{seconds:02d}min"
                else:
                    duration = f"{minutes}min"
            else:
                duration = str(step.end_condition_value)
        elif step.end_condition == 'distance':
            # Formatta la distanza
            if isinstance(step.end_condition_value, (int, float)):
                if step.end_condition_value >= 1000:
                    duration = f"{step.end_condition_value / 1000:.1f}km".replace('.0', '')
                else:
                    duration = f"{int(step.end_condition_value)}m"
            else:
                duration = str(step.end_condition_value)
        
        # Target
        target_text = ""
        if step.target and step.target.target != 'no.target':
            if hasattr(step.target, 'target_zone_name') and step.target.target_zone_name:
                target_text = f" @ {step.target.target_zone_name}"
            else:
                # Determina il tipo di target
                if step.target.target == 'heart.rate.zone' and step.target.from_value and step.target.to_value:
                    from_value = int(step.target.from_value)
                    to_value = int(step.target.to_value)
                    if from_value == to_value:
                        target_text = f" @ {from_value} bpm"
                    else:
                        target_text = f" @ {from_value}-{to_value} bpm"
                
                elif step.target.target == 'pace.zone' and step.target.from_value and step.target.to_value:
                    # Converti da m/s a min/km
                    from_pace_secs = int(1000 / step.target.from_value) if step.target.from_value > 0 else 0
                    to_pace_secs = int(1000 / step.target.to_value) if step.target.to_value > 0 else 0
                    
                    from_pace_mins = from_pace_secs // 60
                    from_pace_s = from_pace_secs % 60
                    
                    to_pace_mins = to_pace_secs // 60
                    to_pace_s = to_pace_secs % 60
                    
                    from_pace = f"{from_pace_mins}:{from_pace_s:02d}"
                    to_pace = f"{to_pace_mins}:{to_pace_s:02d}"
                    
                    # Se i passi sono uguali, è un passo singolo
                    if from_pace == to_pace:
                        target_text = f" @ {from_pace}"
                    else:
                        target_text = f" @ {from_pace}-{to_pace}"
                
                elif step.target.target == 'power.zone' and step.target.from_value and step.target.to_value:
                    from_value = int(step.target.from_value)
                    to_value = int(step.target.to_value)
                    
                    if from_value == 0 and to_value > 0:
                        target_text = f" @ <{to_value}W"
                    elif from_value > 0 and to_value >= 9999:
                        target_text = f" @ {from_value}W+"
                    elif from_value == to_value:
                        target_text = f" @ {from_value}W"
                    else:
                        target_text = f" @ {from_value}-{to_value}W"
        
        # Descrizione
        description_text = ""
        if step.description:
            description_text = f" -- {step.description}"
        
        # Combina tutto
        return f"{indent}{step_type}: {duration}{target_text}{description_text}"
    
    @staticmethod
    def add_steps_to_dataframe(df: 'pd.DataFrame', steps: List[WorkoutStep], start_row: int, indent: str = '') -> int:
        """
        Aggiunge step a un dataframe.
        
        Args:
            df: Dataframe da popolare
            steps: Lista di step da aggiungere
            start_row: Indice di riga iniziale
            indent: Indentazione per gli step nidificati
            
        Returns:
            Indice della prossima riga libera
        """
        row_idx = start_row
        
        # Ottieni la configurazione per convertire i valori in nomi di zone
        config = get_config()
        
        # Per ogni step
        for step in steps:
            # Caso speciale: repeat
            if step.step_type == 'repeat':
                # Aggiungi la riga repeat
                df.loc[row_idx] = {
                    'Nome': '',
                    'Sport': '',
                    'Data': '',
                    'Tipo': f"{indent}repeat",
                    'Condizione di fine': 'iterations',
                    'Valore': step.end_condition_value,
                    'Descrizione': '',
                    'Target': ''
                }
                
                row_idx += 1
                
                # Aggiungi gli step figli con indentazione
                row_idx = ExcelService.add_steps_to_dataframe(df, step.workout_steps, row_idx, indent + '  ')
                
                # Aggiungi la riga end_repeat
                df.loc[row_idx] = {
                    'Nome': '',
                    'Sport': '',
                    'Data': '',
                    'Tipo': f"{indent}end_repeat",
                    'Condizione di fine': '',
                    'Valore': '',
                    'Descrizione': '',
                    'Target': ''
                }
                
                row_idx += 1
                
            else:
                # Step normale
                # Formatta il valore della condizione di fine
                end_value = ''
                
                if step.end_condition == 'lap.button':
                    end_value = 'lap-button'
                elif step.end_condition == 'time':
                    # Formatta il tempo
                    if isinstance(step.end_condition_value, str) and ":" in step.end_condition_value:
                        # Già nel formato mm:ss
                        end_value = step.end_condition_value
                    elif isinstance(step.end_condition_value, (int, float)):
                        # Converti secondi in mm:ss
                        seconds = int(step.end_condition_value)
                        minutes = seconds // 60
                        seconds = seconds % 60
                        end_value = f"{minutes}:{seconds:02d}"
                    else:
                        end_value = str(step.end_condition_value)
                elif step.end_condition == 'distance':
                    # Formatta la distanza
                    if isinstance(step.end_condition_value, str):
                        end_value = step.end_condition_value
                    elif isinstance(step.end_condition_value, (int, float)):
                        # Converti metri in m o km
                        if step.end_condition_value >= 1000:
                            end_value = f"{step.end_condition_value / 1000:.2f}km".replace('.00', '')
                        else:
                            end_value = f"{int(step.end_condition_value)}m"
                    else:
                        end_value = str(step.end_condition_value)
                else:
                    end_value = str(step.end_condition_value) if step.end_condition_value is not None else ''
                
                # Formatta il target
                target_str = ''
                
                if step.target and step.target.target != 'no.target':
                    # Converti valori numerici in nomi di zone
                    if step.target.target == 'pace.zone' and step.target.from_value and step.target.to_value:
                        # Cerca la zona di passo corrispondente
                        paces = config.get(f'sports.{step.sport_type if hasattr(step, "sport_type") else "running"}.paces', {})
                        zone_name = None
                        
                        # Converti da m/s a min/km
                        min_pace_secs = int(1000 / step.target.from_value)
                        max_pace_secs = int(1000 / step.target.to_value)
                        
                        min_pace = f"{min_pace_secs // 60}:{min_pace_secs % 60:02d}"
                        max_pace = f"{max_pace_secs // 60}:{max_pace_secs % 60:02d}"
                        
                        # Cerca la zona corrispondente
                        for name, pace_range in paces.items():
                            if '-' in pace_range:
                                pace_min, pace_max = pace_range.split('-')
                                if pace_min.strip() == min_pace and pace_max.strip() == max_pace:
                                    zone_name = name
                                    break
                            elif pace_range == min_pace and pace_range == max_pace:
                                zone_name = name
                                break
                        
                        if zone_name:
                            target_str = f"@ {zone_name}"
                        else:
                            if min_pace == max_pace:
                                target_str = f"{min_pace}"
                            else:
                                target_str = f"{min_pace}-{max_pace}"
                    
                    elif step.target.target == 'heart.rate.zone' and step.target.from_value and step.target.to_value:
                        # Cerca la zona HR corrispondente
                        heart_rates = config.get('heart_rates', {})
                        zone_name = None
                        
                        # Valori numerici della zona
                        from_value = int(step.target.from_value)
                        to_value = int(step.target.to_value)
                        
                        # Cerca la zona corrispondente
                        for name, hr_range in heart_rates.items():
                            if name.endswith('_HR'):
                                # Calcola valori effettivi usando max_hr
                                max_hr = heart_rates.get('max_hr', 180)
                                
                                if '-' in hr_range and 'max_hr' in hr_range:
                                    # Formato: 62-76% max_hr
                                    parts = hr_range.split('-')
                                    min_percent = float(parts[0])
                                    max_percent = float(parts[1].split('%')[0])
                                    hr_min = int(min_percent * max_hr / 100)
                                    hr_max = int(max_percent * max_hr / 100)
                                    
                                    if hr_min <= from_value <= hr_max and hr_min <= to_value <= hr_max:
                                        zone_name = name
                                        break
                        
                        if zone_name:
                            target_str = f"@ {zone_name}"
                        else:
                            target_str = f"{from_value}-{to_value} bpm"
                    
                    elif step.target.target == 'power.zone' and step.target.from_value and step.target.to_value:
                        # Cerca la zona di potenza corrispondente
                        power_values = config.get('sports.cycling.power_values', {})
                        zone_name = None
                        
                        # Valori numerici della zona
                        from_value = int(step.target.from_value)
                        to_value = int(step.target.to_value)
                        
                        # Cerca la zona corrispondente
                        for name, power_range in power_values.items():
                            if '-' in power_range:
                                power_min, power_max = power_range.split('-')
                                if int(power_min) == from_value and int(power_max) == to_value:
                                    zone_name = name
                                    break
                            elif power_range.startswith('<'):
                                power_val = int(power_range[1:])
                                if from_value == 0 and to_value == power_val:
                                    zone_name = name
                                    break
                            elif power_range.endswith('+'):
                                power_val = int(power_range[:-1])
                                if from_value == power_val and to_value == 9999:
                                    zone_name = name
                                    break
                            else:
                                power_val = int(power_range)
                                if from_value == power_val and to_value == power_val:
                                    zone_name = name
                                    break
                        
                        if zone_name:
                            target_str = f"@ {zone_name}"
                        else:
                            target_str = f"{from_value}-{to_value} W"
                
                # Aggiungi lo step al dataframe
                df.loc[row_idx] = {
                    'Nome': '',
                    'Sport': '',
                    'Data': '',
                    'Tipo': f"{indent}{step.step_type}",
                    'Condizione di fine': step.end_condition,
                    'Valore': end_value,
                    'Descrizione': step.description,
                    'Target': target_str
                }
                
                row_idx += 1
        
        return row_idx